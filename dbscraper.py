# -*- coding: utf-8 -*-
"""
Created on Mon Feb 25 19:08:36 2019

@author: Sandesh Jain
"""

#import all libraries for html interfacing & article reading

import requests 
from bs4 import BeautifulSoup
from newspaper import Article 
from selenium import webdriver

stk = 'aapl '

#Below we take individual inputs from user and form a Google advanced search url
def inputs(Q, St, Si, Yr):
    s1 = 'https://www.google.com/search?as_q='
    quat = Q
    yr= str(Yr)
    s2 = '&as_epq='
    stock = St
    s3='&as_oq=&as_eq=&as_nlo=&as_nhi=&lr=&cr=&as_qdr=all&as_sitesearch='
    site = Si
    s4 = '&as_occt=any&safe=images&as_filetype=&as_rights='
    #url = s1+stock+quat+yr+s2+Q+s3+site+s4
    url = s1+stock+'+'+quat+'+'+yr+s2+Q+s3+site+s4
    #print url
    return url

#Here we slice the url received from inputs() function
def google_adv_search(z1):
    r=requests.get(z1)
    soup=BeautifulSoup(r.text, 'html.parser')
    results=soup.find('div', attrs={'id':'search'})
    all=results.find('a')['href']
    all = all[7:]
    all = all.split('&sa')[0]
    return all

#Below we retrieve stock name directly from yahoo finance web service as json and return the stock company name
def get_symbol(symbol):
    url = "http://d.yimg.com/autoc.finance.yahoo.com/autoc?query={}&region=1&lang=en".format(symbol)
    result = requests.get(url).json()
    for x in result['ResultSet']['Result']:
        if x['symbol'] == symbol:
            return x['name']


# Open the link found by Google and extract the Outputs - title, author, date, text, symbol name & link
def read_website(z2):
    search = z2
    #print search
    article = Article(search) #using the newsaper library's Article function
    article.download()
    article.html
    article.parse()
    db=[]
    db.append(article.title)
    db.append(stk)
    db.append(article.publish_date)
    db.append(search)
    # Get stock symbol for inclusion in excel
    stock = stk
    company = get_symbol(stock.upper())
    db.append(company)
    db.append(article.text)
    return db

# Using openpyexcel populate sheet by the output data
def write_to_excel(z3, n):
    # Import `load_workbook` module from `openpyxl`
    from openpyxl import load_workbook
    wb = load_workbook('./stocknow.xlsx') # Make an excel worksheet with name stocknow
    sheet = wb.get_sheet_by_name('Sheet1')
    db = z3
    sheet['A'+str(n)] = db[0]  #title
    sheet['B'+str(n)] = db[1]  #author
    sheet['C'+str(n)] = db[2]  #date of publication
    sheet['D'+str(n)] = db[3]  #the link
    sheet['E'+str(n)] = db[4]  #company name
    sheet['F'+str(n)] = db[5]  # the article
    
    wb.save('./stocknow.xlsx')
browser = webdriver.Firefox(executable_path=r'C:\Users\Administrator\Downloads\geckodriver-v0.24.0-win32\geckodriver.exe')
t = 0
n=85    
stk = 'GM '
import time
Quarters = ['First-Quarter' ,'Second-Quarter','Third-Quarter','Fourth-Quarter']
Quart = ['q1+','q2+','q3+','q4+']
#Quart = ['q3 ','q4 ']
shares = ['X+', 'BGS+', 'BP+']
for stk in shares:
    for q in Quart:
        Q = q
        St = stk
        Si = 'seekingalpha.com'
        Yrs = [2016,2017,2018]
        for Yr in Yrs:
            browser = webdriver.Firefox(executable_path=r'C:\Users\Administrator\Downloads\geckodriver-v0.24.0-win32\geckodriver.exe')
            n = n + 1
            z1 =  inputs(Q, St, Si, Yr)
            print z1
            z2 =  google_adv_search(z1)
            z3 =  browser.get(z2)
            j=browser.find_elements_by_xpath('//*[@id="a-body"]')
            i = ''
            i1=''
            i2=''
            for elem in j:
                #print elem.text
                i = i + elem.text.encode('utf-8')
            j1 = browser.find_elements_by_xpath('//*[@id="a-hd"]/h1')
            for elem in j1:
                #print elem.text
                i1 = i1 + elem.text.encode('utf-8')
            j2 = browser.find_elements_by_xpath('//*[@id="a-hd"]/div[2]/time')    
            for elem in j2:
                #print elem.text
                i2 = i2 + elem.text.encode('utf-8')
             # Import `load_workbook` module from `openpyxl`
            #print i1
            #print i2
            #print i
            time.sleep(0.2)    
    
            from openpyxl import load_workbook
            wb = load_workbook('./stocknow.xlsx') # Make an excel worksheet with name stocknow
            sheet = wb.get_sheet_by_name('Sheet1')    
            sheet['A'+str(n)] = i1    
            sheet['B'+str(n)] = stk
            sheet['C'+str(n)] = i2  
            sheet['D'+str(n)] = z2
            sheet['E'+str(n)] = get_symbol(stk.upper())
            sheet['F'+str(n)] = i
            wb.save('./stocknow.xlsx')  
            
            browser.quit()
            #write_to_excel(z3 , n)
            t = t + 1
            print('Done'+str(t))

#browser.get('https://seekingalpha.com/article/4090041-alphabets-goog-ceo-sundar-pichai-q2-2017-results-earnings-call-transcript')
#j=browser.find_elements_by_xpath('//*[@id="a-body"]')
#for elem in j:
#    print elem.text
#browser.quit()
            # Import `load_workbook` module from `openpyxl`
#        from openpyxl import load_workbook
#        wb = load_workbook('./stocknow.xlsx') # Make an excel worksheet with name stocknow
#        sheet = wb.get_sheet_by_name('Sheet1')
#        db = z3
#        sheet['A'+str(n)] = browser.find_elements_by_xpath('//*[@id="a-hd"]/h1').text  #title
#        sheet['B'+str(n)] = 'Transcript'  #author
#        sheet['C'+str(n)] = browser.find_elements_by_xpath('//*[@id="a-hd"]/div[2]/time').text  #date of publication
#        sheet['D'+str(n)] = z2  #the link
#        sheet['E'+str(n)] = browser.find_elements_by_xpath('//*[@id="a-body"]/p[1]/a').text  #company name
#        sheet['F'+str(n)] = i  # the article
#    
#        wb.save('./stocknow.xlsx')  

















































# -*- coding: utf-8 -*-
"""
Created on Wed Feb 13 17:44:03 2019

@author: Administrator
"""



# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load the workbook in wb
wb = load_workbook('./stocknow.xlsx')

# load the Sheet1 in sheet
sheet = wb.get_sheet_by_name('Sheet1')

# edit the sheet's cell value
sheet['A1'] = 'fake error'

# close the modified xlsx file the save it using
wb.save('./stocknow.xlsx')

import webbrowser
webbrowser.open('http://inventwithpython.com/')
